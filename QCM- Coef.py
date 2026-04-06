from pathlib import Path
from openpyxl import load_workbook, Workbook


def calculer_scores_participant(grille_cotation, ligne_participant):
    """
    Calcule les scores d'un participant à partir d'une ligne de réponses Excel.
    """

    identifiant_participant = ligne_participant[1]

    scores = {
        "Résilience": 0,
        "Esprit critique": 0,
        "Comportement social": 0,
        "Compétences techniques": 0,
        "Traitement de l'information": 0,
        "Création": 0,
    }

    correspondance_dimensions = {
        "Résilience": "Résilience",
        "EC": "Esprit critique",
        "CSDLEN": "Comportement social",
        "CT": "Compétences techniques",
        "TDLinfo": "Traitement de l'information",
        "CDC": "Création",
    }

    # Les réponses commencent à partir de la 4e colonne
    for index, reponse_participant in enumerate(ligne_participant[3:]):
        code_dimension = grille_cotation[index][0]
        reponses_possibles = grille_cotation[index][1:]

        if reponse_participant is None:
            continue

        reponse_participant = str(reponse_participant).strip()

        for reponse_attendue, score in reponses_possibles:
            if reponse_participant == reponse_attendue:
                nom_dimension = correspondance_dimensions[code_dimension]
                scores[nom_dimension] += score
                break

    # Ramener les scores négatifs à 0
    for dimension in scores:
        if scores[dimension] < 0:
            scores[dimension] = 0

    score_total = sum(scores.values())

    return [
        identifiant_participant,
        scores["Résilience"],
        scores["Esprit critique"],
        scores["Comportement social"],
        scores["Compétences techniques"],
        scores["Traitement de l'information"],
        scores["Création"],
        score_total,
    ]


# Grille de cotation
GRILLE_COTATION = [
    ('Résilience', ('A', 0.5), ('B', 1), ('C', -1), ('D', -0.5), ('E', 0)),
    ('Résilience', ('A', -0.5), ('B', -1), ('C', 1), ('D', -0.5), ('E', 0)),
    ('EC', ('A', -1), ('B', 1), ('C', -0.5), ('D', -0.5), ('E', 0)),
    ('CSDLEN', ('A', -0.5), ('B', 1), ('C', -0.5), ('D', -0.5), ('E', 0)),
    ('CSDLEN', ('A', -0.5), ('B', 1), ('C', -1), ('D', -0.5), ('E', 0)),
    ('CDC', ('A', -0.5), ('B', -0.5), ('C', 1), ('D', -1), ('E', 0)),
    ('Résilience', ('A', 1), ('B', 0), ('C', 0.5), ('D', -1), ('E', 0)),
    ('CDC', ('A', -1), ('B', -0.5), ('C', 1), ('D', -1), ('E', 0)),
    ('CT', ('1', 1), ('2', -1), ('3', -0.5), ('4', -1), ('5', 0)),
    ('TDLinfo', ('A', 0.5), ('B', -1), ('C', 1), ('D', 0), ('E', 0)),
    ('CDC', ('A', -1), ('B', 1), ('C', 1), ('D', -1), ('E', 0)),
    ('CT', ('1', 1), ('2', -0.5), ('3', -1), ('4', -1), ('5', 0)),
    ('CDC', ('A', -1), ('B', 0.5), ('C', 1), ('D', -0.5), ('E', 0)),
    ('Résilience', ('A', -0.5), ('B', 0.5), ('C', -1), ('D', 1), ('E', 0)),
    ('CT', ('1', 0.5), ('2', 1), ('3', -1), ('4', -1), ('5', 0)),
    ('CSDLEN', ('A', -1), ('B', -0.5), ('C', -1), ('D', 1), ('E', 0)),
    ('EC', ('A', 1), ('B', -1), ('C', -0.5), ('D', -1), ('E', 0)),
    ('TDLinfo', ('A', 1), ('B', 0.5), ('C', -1), ('D', -1), ('E', 0)),
    ('TDLinfo', ('A', 1), ('B', -1), ('C', -0.5), ('D', 0.5), ('E', 0)),
    ('CT', ('1', -1), ('2', -1), ('3', 1), ('4', -1), ('5', 0)),
    ('CT', ('1', -1), ('2', 1), ('3', -1), ('4', -1), ('5', 0)),
    ('CDC', ('A', -0.5), ('B', 1), ('C', -1), ('D', -1), ('E', 0)),
    ('TDLinfo', ('A', -1), ('B', -1), ('C', 1), ('D', -0.5), ('E', 0)),
    ('CSDLEN', ('A', -1), ('B', -0.5), ('C', 1), ('D', -1), ('E', 0)),
    ('EC', ('A', -1), ('B', -1), ('C', 1), ('D', -1), ('E', 0)),
    ('CSDLEN', ('A', -0.5), ('B', 1), ('C', -0.5), ('D', -0.5), ('E', 0)),
    ('Résilience', ('A', -0.5), ('B', -1), ('C', 1), ('D', 0), ('E', 0)),
    ('EC', ('A', 1), ('B', -1), ('C', -1), ('D', -0.5), ('E', 0)),
    ('TDLinfo', ('A', 0.5), ('B', -1), ('C', 1), ('D', -0.5), ('E', 0)),
    ('EC', ('A', -1), ('B', 1), ('C', -0.5), ('D', -1), ('E', 0)),
]


def main():
    fichier_entree = Path("donnees/questionnaire_reponses.xlsx")
    fichier_sortie = Path("resultats/scores_questionnaire.xlsx")

    fichier_sortie.parent.mkdir(parents=True, exist_ok=True)

    classeur_entree = load_workbook(fichier_entree)
    feuille_entree = classeur_entree.active

    donnees = [list(row) for row in feuille_entree.iter_rows(values_only=True)]

    # Suppression de la ligne d'en-tête
    donnees = donnees[1:]

    classeur_sortie = Workbook()
    feuille_sortie = classeur_sortie.active
    feuille_sortie.title = "Scores"

    en_tetes = [
        "Participant",
        "Résilience",
        "Esprit critique",
        "Comportement social",
        "Compétences techniques",
        "Traitement de l'information",
        "Création",
        "Total",
    ]
    feuille_sortie.append(en_tetes)

    for ligne_participant in donnees:
        scores = calculer_scores_participant(GRILLE_COTATION, ligne_participant)
        feuille_sortie.append(scores)

    classeur_sortie.save(fichier_sortie)
    print(f"Les résultats ont été enregistrés dans : {fichier_sortie}")


if __name__ == "__main__":
    main()
